import sys
import PyQt5.QtWidgets
from PyQt5 import uic
from PyQt5.QtCore import *
import webbrowser
import os
import hashlib
from PyQt5 import QtCore
from PyQt5 import QtGui
from PyQt5.QtGui import QIcon
import csv
from datetime import datetime
import paramiko
import json
from time import *
from PyQt5.QtTest import *
import pandas as pd
from oauth2client.service_account import ServiceAccountCredentials
import gspread
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Border, Side
from pycaret.classification import *

def resource_path(relative_path):
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)

form = resource_path("GUI.ui")
form_class = uic.loadUiType(form)[0]
form2 = resource_path("Report.ui")
spreadKey = resource_path("CTRD_Upload_Spread.json")
model_load = resource_path("CTRD_Label_Model")
config_load = resource_path("CTRD_Config_Model.pkl")
class Thread(QThread):
    _signal = pyqtSignal(int)
    def __init__(self):
        super(Thread, self).__init__()
    
    def run(self):
        for i in range(101):
            sleep(2)
            self._signal.emit(i)

class OptionWindow(PyQt5.QtWidgets.QDialog):
    def __init__(self, parent):
        super(OptionWindow, self).__init__(parent)  
        self.ui = uic.loadUi(form2, self)         
        self.show()
        self.setWindowTitle('CTRD v1.4 Detection Report')
        self.FilePath.setText(filename[0])
        self.FileSize.setText(filesize)
        self.Hash.setText(sha256)
        self.Result.setText(MLResult)
        self.SaveCSV.clicked.connect(self.CreateReport)

    def CreateReport(self):
        box = Border(bottom=Side(border_style="thick", color='00000000'))

        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'CTRD REPORT'
        ca1 = ws['A1']
        ca1.font = Font(name='Segoe UI Black', size=26)
        ca1.alignment = Alignment(horizontal='center', vertical='center')
        ca1.border = box
        ws.merge_cells('A1:J1')

        ws2 = wb.active
        ws2.merge_cells('A2:B4')
        ws2['A2'] = 'File Path'
        ca2 = ws['A2']
        ca2.font = Font(name='Bahnschrift SemiBold SemiConden', size=12)
        ca2.alignment = Alignment(vertical='center')

        ws3 = wb.active
        ws3.merge_cells('C2:J4')
        ws3['C2'] = filepath
        ca3 = ws['C2']
        ca3.font = Font(name='Bahnschrift SemiLight SemiConde', size=11)
        ca3.alignment = Alignment(vertical='center')

        ws4 = wb.active
        ws4.merge_cells('A5:B7')
        ws4['A5'] = 'File Size'
        ca4 = ws['A5']
        ca4.font = Font(name='Bahnschrift SemiBold SemiConden', size=12)
        ca4.alignment = Alignment(vertical='center')

        ws5 = wb.active
        ws5.merge_cells('C5:J7')
        ws5['C5'] = filesize
        ca5 = ws['C5']
        ca5.font = Font(name='Bahnschrift SemiLight SemiConde', size=11)
        ca5.alignment = Alignment(vertical='center')

        ws6 = wb.active
        ws6.merge_cells('A8:B10')
        ws6['A8'] = 'SHA256 Hash'
        ca6 = ws['A8']
        ca6.font = Font(name='Bahnschrift SemiBold SemiConden', size=12)
        ca6.alignment = Alignment(vertical='center')

        ws7 = wb.active
        ws7.merge_cells('C8:J10')
        ws7['C8'] = sha256
        ca7 = ws['C8']
        ca7.font = Font(name='Bahnschrift SemiLight SemiConde', size=11)
        ca7.alignment = Alignment(vertical='center')

        ws8 = wb.active
        ws8.merge_cells('A11:B13')
        ws8['A11'] = 'Detection Result'
        ca8 = ws['A11']
        ca8.font = Font(name='Bahnschrift SemiBold SemiConden', size=12)
        ca8.alignment = Alignment(vertical='center')

        ws9 = wb.active
        ws9.merge_cells('C11:J13')
        ws9['C11'] = MLResult
        ca9 = ws['C11']
        ca9.font = Font(name='Bahnschrift SemiLight SemiConde', size=11)
        ca9.alignment = Alignment(vertical='center')

        ws10 = wb.active
        ws10['A14'] = 'Scan Date'
        ca10 = ws['A14']
        ca10.font = Font(name='Bahnschrift SemiBold SemiConden', size=12)
        ca10.alignment = Alignment(vertical='center')
        ca10.border = box
        ws10.merge_cells('A14:B16')

        ws11 = wb.active
        ws11['C14'] = now.strftime(f'%Y-%m-%d %H:%M:%S')
        ca11 = ws['C14']
        ca11.font = Font(name='Bahnschrift SemiLight SemiConde', size=11)
        ca11.alignment = Alignment(vertical='center')
        ca11.border = box
        ws11.merge_cells('C14:J16')

        wb.save('CTRD_Report/{0}_CTRD_Report.xlsx' .format(sha256))
        msgBox = PyQt5.QtWidgets.QMessageBox()
        msgBox.setStyleSheet('QMessageBox {color:black; background:white;}')
        msgBox.information(msgBox,'Notice','CTRD 결과보고서 파일이 생성되었습니다.\n\nCTRD_Report 폴더를 확인해 주시기 바랍니다.', msgBox.Ok)

class MyWindow(PyQt5.QtWidgets.QMainWindow, form_class):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('CTRD v1.0')
        self.setWindowFlag(Qt.FramelessWindowHint)
        self.setupUi(self)
        self.Initial.clicked.connect(self.InitialMethod)
        self.Initial.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.DataFolderCreate()
        self.ReportFolderCreate()
        self.Run.clicked.connect(self.Main)
        self.Run.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.shutdown.clicked.connect(PyQt5.QtWidgets.QApplication.instance().quit)
        self.minimize.clicked.connect(self.hideWindow)
        self.github.clicked.connect(lambda: webbrowser.open('https://github.com/not4dog/Cerber-Type-Ransomware-Detection'))
        self.hongiklogo.clicked.connect(lambda: webbrowser.open('https://sejong.hongik.ac.kr/index.do'))
        self.github.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.hongiklogo.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.minimize.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.shutdown.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))

    def InitialMethod(self):
        msgbox = PyQt5.QtWidgets.QMessageBox()
        msgbox.setStyleSheet('QMessageBox {color:black; background:white;}')
        ret = msgbox.question(msgbox,'Question', '초기화 시 기존 분석자료와 결과보고서가 영구적으로 삭제됩니다.\n\n그래도 초기화 하시겠습니까?', msgbox.Yes | msgbox.No)
        if ret == msgbox.Yes:
           QtCore.QCoreApplication.quit()
           QtCore.QProcess.startDetached(sys.executable, sys.argv)
           os.system('rmdir /s /q CTRD_Feature_Data & rmdir /s /q CTRD_Report')
        else : return

    def pBar(self):
        self.thread = Thread()
        self.thread._signal.connect(self.signal_accept)
        self.thread.start()

    def DataFolderCreate(self):
        dir_path = "CTRD_Feature_Data"
        if os.path.isdir(dir_path) != True :
            os.system('mkdir CTRD_Feature_Data')
        else : return

    def ReportFolderCreate(self):
        dir_path = "CTRD_Report"
        if os.path.isdir(dir_path) != True :
            os.system('mkdir CTRD_Report')
        else : return

    def signal_accept(self, msg):
        self.progressBar.setValue(int(msg))
        if self.progressBar.value() == 100:
            self.progressBar.setValue(0)
            self.msg_box()
        return

    def msg_box(self):
        msg = PyQt5.QtWidgets.QMessageBox()
        ret = msg.information(msg,'Notice', '실행파일 분석이 완료되었습니다.\n\nOK 버튼을 클릭하시면 CTRD 결과보고서 확인이 가능합니다.', msg.Ok | msg.Cancel)
        if ret == msg.Ok:
           OptionWindow(self)
           self.Run.setDisabled(False)
        else : self.Run.setDisabled(False)
        return

    def hideWindow(self):
        self.showMinimized()
    
    def CheckSHA256(self):
        global sha256
        hash_sha256 = hashlib.sha256()
        with open(filename[0], "rb") as f:
            chunk = f.read(4096)
            while chunk:
                hash_sha256.update(chunk)
                chunk = f.read(4096)
        sha256 = hash_sha256.hexdigest()
        return

    def CountAPI(self, item):
        file = open(f"CTRD_Feature_Data\{sha256}_API_Extract.json", "r")
        read_data = file.read()
        word_count = read_data.lower().count(item)
        return word_count

    def ExtractOpcode(self) :
        self.CheckSHA256()
        os.system('objdump -d -j .text {0} > CTRD_Feature_Data\{1}_Opcode_Extract.txt' .format(filename[0], sha256))
        push = self.CountOpcode("push")
        mov = self.CountOpcode("mov")
        call = self.CountOpcode("call")
        sub = self.CountOpcode("sub")
        jmp = self.CountOpcode("jmp")
        add = self.CountOpcode("add")
        cmp = self.CountOpcode("cmp")
        test = self.CountOpcode("test")
        lea = self.CountOpcode("lea")
        pop = self.CountOpcode("pop")

        f = open(f'CTRD_Feature_Data\{sha256}_Opcode_Frequency.csv','w', newline='')
        wr = csv.writer(f)
        wr.writerow(["SHA256", "push", "mov", "call", "sub", "jmp", "add", "cmp", "test", "lea", "pop"])
        wr.writerow([sha256, push, mov, call, sub, jmp, add, cmp, test, lea, pop])
        f.close()
        return

    def CountOpcode(self, item):
        file = open(f"CTRD_Feature_Data\{sha256}_Opcode_Extract.txt", "r")
        read_data = file.read()
        word_count = read_data.lower().count(item)
        return word_count

    def FileSize(self, size_bytes):
        import math
        if size_bytes == 0:
            return "0B"
        size_name = ("B", "KB", "MB", "GB", "TB", "PB", "EB", "ZB", "YB")
        i = int(math.floor(math.log(size_bytes, 1024)))
        p = math.pow(1024, i)
        s = round(size_bytes / p, 2)
        return "%s %s" % (s, size_name[i])
    
    def reset(self):
        loop = QEventLoop()
        QTimer.singleShot(5000, loop.quit) 
        loop.exec_()

    def sshConnect(self):
        global ssh
        PyQt5.QtWidgets.QApplication.processEvents()
        ssh = paramiko.SSHClient()
        PyQt5.QtWidgets.QApplication.processEvents()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        PyQt5.QtWidgets.QApplication.processEvents()
        ssh.connect('211.214.61.14',port='2200',username='b793170',password ='20100709')
        PyQt5.QtWidgets.QApplication.processEvents()

    def FileTransper(self):
        global sftp
        sftp =ssh.open_sftp()
        PyQt5.QtWidgets.QApplication.processEvents()
        remotepath = '/home/b793170/Desktop/Scan.exe' 
        localpath  = filepath 
        PyQt5.QtWidgets.QApplication.processEvents()
        sftp.put(localpath, remotepath)
        PyQt5.QtWidgets.QApplication.processEvents()

    def Analysis(self):
        stdin, stdout, stderr = ssh.exec_command('curl -H "Authorization: Bearer pxJLRqiTfxz0PNNhGLdoew" -F file=@/home/b793170/Desktop/Scan.exe http://localhost:8090/tasks/create/file')
        PyQt5.QtWidgets.QApplication.processEvents()

    def Exists(self):
        output = False
        result = False

        while True :
            stdin, stdout, stderr = ssh.exec_command('[ -f /home/b793170/.cuckoo/storage/analyses/1/reports/report.json ] && echo "$FILE True" || echo "$FILE False"')
            output =''.join(stdout.readlines())
            result = output.replace(" ","")
            json.loads(result.lower())
            PyQt5.QtWidgets.QApplication.processEvents()
            self.reset()
            PyQt5.QtWidgets.QApplication.processEvents()
            if json.loads(result.lower()) != False :
                break
            PyQt5.QtWidgets.QApplication.processEvents()

    def FileTransperAndExtract(self):
        remotepath2 = '/home/b793170/.cuckoo/storage/analyses/1/reports/report.json'
        localpath2 = 'CTRD_Feature_Data\{0}_API_Extract.json' .format(sha256)
        sftp.get(remotepath2, localpath2)
        PyQt5.QtWidgets.QApplication.processEvents()
        stdin, stdout, stderr = ssh.exec_command("rm -f /home/b793170/Desktop/Scan.exe")
        PyQt5.QtWidgets.QApplication.processEvents()
        stdin, stdout, stderr = ssh.exec_command('curl -H "Authorization: Bearer pxJLRqiTfxz0PNNhGLdoew" http://localhost:8090/tasks/delete/1')
        PyQt5.QtWidgets.QApplication.processEvents()
        ssh.close()
        PyQt5.QtWidgets.QApplication.processEvents()
        sftp.close()
        PyQt5.QtWidgets.QApplication.processEvents()

        api1 = self.CountAPI("findfirstfile")
        api2 = self.CountAPI("searchpathw")
        api3 = self.CountAPI("setfilepointer")
        api4 = self.CountAPI("findresourceex")
        api5 = self.CountAPI("getfileattributesw")
        api6 = self.CountAPI("setfileattributesw")
        api7 = self.CountAPI("setfilepointerex")
        api8 = self.CountAPI("cryptencrypt")
        api9 = self.CountAPI("createthread")
        api10 = self.CountAPI("findresourceexw")

        f = open(f'CTRD_Feature_Data\{sha256}_API_Frequency.csv','w', newline='' .format(sha256))
        wr = csv.writer(f)
        wr.writerow(["FindFirstFile", "SearchPathW", "SetFilePointer", "FindResourceEx", "GetFileAttributesW", "SetFileAttributesW", "SetFilePointerEx", "CryptEncrypt", "CreateThread", "FindResourceExW"])
        wr.writerow([api1, api2, api3, api4, api5, api6, api7, api8, api9, api10])
        f.close()
        return

    def sleep(self):
        QTest.qWait(5000)

    def FeatureMerge(self):
        Opcode = 'CTRD_Feature_Data/{0}_Opcode_Frequency.csv' .format(sha256)
        api = 'CTRD_Feature_Data/{0}_API_Frequency.csv' .format(sha256)

        dataFrame = pd.concat(map(pd.read_csv, [Opcode, api]), axis=1)
        dataFrame.to_csv('CTRD_Feature_Data/{0}_All_Feature_CTRD_Data.csv' .format(sha256), index = False)

    def UploadSpread(self):
        scope = ["https://spreadsheets.google.com/feeds",
                "https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive.file",
                "https://www.googleapis.com/auth/drive"]

        creds = ServiceAccountCredentials.from_json_keyfile_name(spreadKey, scope)
        PyQt5.QtWidgets.QApplication.processEvents()

        spreadsheet_name = "CTRD_Feature_Data"
        client = gspread.authorize(creds)
        PyQt5.QtWidgets.QApplication.processEvents()
        spreadsheet = client.open(spreadsheet_name)
        PyQt5.QtWidgets.QApplication.processEvents()

        for sheet in spreadsheet.worksheets():
            PyQt5.QtWidgets.QApplication.processEvents()
            sheet

        new_df = pd.read_csv('CTRD_Feature_Data/{0}_All_Feature_CTRD_Data.csv'.format(sha256))
        PyQt5.QtWidgets.QApplication.processEvents()
        val_list = new_df.values.tolist()
        load_list =val_list[0]

        sheet.append_row(load_list)
        PyQt5.QtWidgets.QApplication.processEvents()

    def CTRD_ML(self):
        model = load_model(model_load)
        upload_data= pd.read_csv('CTRD_Feature_Data/{0}_All_Feature_CTRD_Data.csv' .format(sha256))

        load_config(config_load)
        prep_pipe = get_config('prep_pipe')

        Score = prep_pipe.predict_proba(upload_data)
        Convert_Score = Score[0,:]
        Convert_Benign_Score = round(Convert_Score[0] * 100, 2)
        Convert_Cerber_Score = round(Convert_Score[1] * 100, 2)

        global MLResult

        if Convert_Cerber_Score > 50.00 :
            os.remove(filepath)
        else : pass

        MLResult = "This File has an {0}% Chance of being CERBER-TYPE RANSOMWARE." .format(Convert_Cerber_Score)
        return

    def Main(self):
        global filename, filesize
        global filepath
        global now 
        now = datetime.now()
        filename = PyQt5.QtWidgets.QFileDialog.getOpenFileName(self, 'Choose Executable File', 'C:/', 'Executable File (*.exe)')
        filepath = filename[0]

        if filename[0] !='' :
            with open(filename[0], 'rb') as f:
                signature1 = f.read(4)
                signature2 = signature1
    
            if signature1 == b'MZ\x90\x00' or signature2 == b'MZP\x00' :
               pass

            else :
               msgBox = PyQt5.QtWidgets.QMessageBox()
               msgBox.setStyleSheet('QMessageBox {color:black; background:white;}')
               msgBox.warning(msgBox,'Warning','선택한 파일은 실행파일이 아닙니다.\n\n올바른 실행파일을 선택해 주시기 바랍니다.')
               return(print('실행파일이 아닌 파일 선택으로 인한 메인함수 중단'))
               
            file_size = os.path.getsize('{0}' .format(filename[0]))
            filesize = self.FileSize(file_size)

            self.progressBar.setFormat("Processing...")
            self.pBar()
            self.Run.setDisabled(True)
            self.ExtractOpcode() 
            self.sshConnect()
            self.FileTransper()
            self.Analysis()
            self.Exists()
            self.FileTransperAndExtract()
            self.FeatureMerge()
            self.UploadSpread()
            self.CTRD_ML()

        else :
            msgBox = PyQt5.QtWidgets.QMessageBox()
            msgBox.setStyleSheet('QMessageBox {color:black; background:white;}')
            msgBox.warning(msgBox,'Warning','분석 대상 파일이 선택되지 않았습니다.\n\n파일을 선택해 주시기 바랍니다.')
            return(print('파일 미 선택으로 인한 메인함수 중단'))

if __name__ == "__main__":
    app = PyQt5.QtWidgets.QApplication(sys.argv)
    app.setWindowIcon(QIcon('icon.ico'))
    myWindow = MyWindow()
    myWindow.show()
    app.exec_()